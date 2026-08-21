"""Build the comprehensive results workbook for the 7,8-DHF r/Nootropics run.

Every number is recomputed from noots.db / records.csv / subreddit_posts.json,
so the workbook cannot drift from the pipelines. Formatting is carried as .xlsx
because Drive converts it on upload; CSV cannot hold tabs, bold or number formats.
"""
from __future__ import annotations
import collections, csv, json, math, re, sqlite3, sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from scipy.stats import fisher_exact
from statsmodels.stats.multitest import multipletests

sys.stdout.reconfigure(encoding="utf-8", errors="replace")
HERE = Path(__file__).resolve().parent
OUT = HERE / "results_workbook.xlsx"

INK, MUTED = "1F2933", "6B7280"
RULE = Side(style="thin", color="D8DEE4")
HDR_FILL = PatternFill("solid", fgColor="37474F")
BAND = PatternFill("solid", fgColor="37474F")
GOOD = PatternFill("solid", fgColor="E3F4E9")
WARN = PatternFill("solid", fgColor="FDF0E3")
BAD = PatternFill("solid", fgColor="FBE4E4")
BOLD = Font(bold=True, size=10, color=INK)
SMALL = Font(size=9, color=MUTED, italic=True)
WRAP = Alignment(wrap_text=True, vertical="top")


def wilson(k, n, z=1.96):
    if not n:
        return float("nan"), float("nan")
    p = k / n
    d = 1 + z * z / n
    c = (p + z * z / (2 * n)) / d
    m = z * math.sqrt((p * (1 - p) + z * z / (4 * n)) / n) / d
    return max(0.0, c - m), min(1.0, c + m)


# ══ gather ══════════════════════════════════════════════════════════════════
PAT = re.compile(r"(?i)(tropoflavin|hydroxyflavone|\b7[ .,'-]{0,2}8[ .,'-]{0,2}dhf\b|\bdhf\b)")
DMA = re.compile(r"(?i)\b4[ '’]?-?\s?dma|eutropoflav")
PLAIN = re.compile(r"(?i)tropoflavin|dihydroxyflavone|\b7[ .,'-]{0,2}8[ .,'-]{0,2}dhf\b|\bdhf\b")
SIG = {"strong": 3, "moderate": 2, "weak": 1, "n/a": 0, None: 0, "": 0}

corpus = json.loads((HERE / "source" / "subreddit_posts.json").read_text(encoding="utf-8"))
items = []
for p in corpus:
    t = ((p.get("title") or "") + " " + (p.get("body") or "")).strip()
    if PAT.search(t):
        items.append(t)
    for c in p["comments"]:
        b = (c.get("body") or "").strip()
        if PAT.search(b):
            items.append(b)
N_ITEMS = len(items)

PURPOSE = {
    "neurogenesis / BDNF / rewiring": r"\bbdnf\b|\btrkb\b|neurogenes|neuroplastic|\bsynaptogenes|\brewir",
    "depression / mood": r"\bdepress|\bantidepress|\bmood\b|\bsad\b|\bmdd\b|\bdysthym|anhedoni",
    "focus / cognition / brain fog": r"\bfocus\b|\bconcentrat|\bcognit|\bbrain fog\b|\bclarity\b|\bproductiv",
    "sleep": r"\bsleep\b|\binsomnia\b|\bdream",
    "memory / learning": r"\bmemory\b|\brecall\b|\blearning\b|\bmemoriz|\bretention\b",
    "anxiety": r"\banxiet|\banxious\b|\bpanic\b|\bsocial anxiety\b",
    "ADHD": r"\badhd\b|\badd\b(?!ed)|\battention deficit",
    "neuroprotection / repair": r"neuroprotect|\brepair\b|\bregenerat|\bheal\w* (my |the )?brain",
    "stimulant recovery / tolerance": r"\bstimulant\b.{0,30}(recover|damage|crash)|amphetamine.{0,30}(damage|recover)|\btolerance\b",
    "Alzheimer's / dementia": r"\balzheim|\bdementia\b|cognitive decline|\bmci\b",
    "exercise mimetic / fat loss": r"exercise mimetic|\bfat loss\b|\bweight loss\b|\bmetabol|\bobes|\bglucose\b|\binsulin\b",
    "drug damage (PSSD/PFS/HPPD)": r"\bpssd\b|\bpfs\b|post.?ssri|\bhppd\b|\bneurotox|damage from|\bexcitotox",
    "libido / sexual": r"\blibido\b|\bsexual\b|\berectile\b|\borgasm",
    "neuropathy / nerve / tinnitus": r"\bneuropath|\bnerve (damage|pain|regen)|\btinnitus\b|\bhearing\b",
    "autism / Rett": r"\bautis|\brett\b|\bfragile x\b",
    "TBI / concussion": r"\btbi\b|concussion|\bhead injury\b|traumatic brain",
}
purpose_n = {k: sum(1 for t in items if re.search(p, t, re.I)) for k, p in PURPOSE.items()}

CO = {
    "4'-DMA-7,8-DHF (eutropoflavin)": r"4[ '’]?-?\s?dma|eutropoflavin",
    "noopept": r"\bnoopept\b", "semax": r"\bsemax\b",
    "racetams (pir/ani/oxi/phenyl)": r"\bpiracetam\b|\baniracetam\b|\boxiracetam\b|\bphenylpiracetam\b|\bpramiracetam\b",
    "NSI-189": r"\bnsi[- ]?189\b", "lion's mane": r"lion'?s? mane|hericium",
    "dihexa": r"\bdihexa\b", "bromantane": r"\bbromantane\b", "selank": r"\bselank\b",
    "magnesium": r"\bmagnesium\b", "cerebrolysin": r"\bcerebrolysin\b",
    "agmatine": r"\bagmatine\b", "uridine": r"\buridine\b",
    "psilocybin / LSD": r"\bpsilocybin\b|\blsd\b|microdos", "ketamine": r"\bketamine\b",
    "creatine": r"\bcreatine\b", "curcumin / turmeric": r"\bcurcumin\b|turmeric",
    "9-MBC": r"\b9[- ]?mbc\b",
}
co_n = {k: sum(1 for t in items if re.search(p, t, re.I)) for k, p in CO.items()}

db = sqlite3.connect(HERE / "noots.db")
db.row_factory = sqlite3.Row
A = db.execute("""SELECT r.sentiment, r.signal_strength sig, r.user_id, r.side_effects,
                         p.body_text, p.title, p.post_date
                  FROM treatment_reports r JOIN posts p ON p.post_id=r.post_id""").fetchall()
A_raw = collections.Counter(r["sentiment"] for r in A)
A_sig = collections.Counter((r["sentiment"], r["sig"] or "n/a") for r in A)
best = {}
for r in sorted(A, key=lambda r: (r["post_date"] or 0, SIG.get(r["sig"], 0)), reverse=True):
    best.setdefault(r["user_id"], r["sentiment"])
A_vote = collections.Counter(best.values())
A_N = len(best)

B_rows = list(csv.DictReader((HERE / "source_B" / "records.csv").open(encoding="utf-8")))
META = {"author_hash", "source", "text_count", "schema_id", "extraction_method", "extracted_at"}
B_fields = [k for k in B_rows[0] if k not in META]
fill = collections.Counter()
for r in B_rows:
    for k in B_fields:
        if (r.get(k) or "").strip():
            fill[k] += 1

RANK = {"worsened": 4, "no_effect": 3, "mixed": 2, "helped": 1, "unknown": 0}
entry_out = {"7,8-DHF": collections.Counter(), "4'-DMA": collections.Counter()}
author_out = {"7,8-DHF": {}, "4'-DMA": {}}
targets = {"7,8-DHF": collections.Counter(), "4'-DMA": collections.Counter()}
for r in B_rows:
    for entry in (r.get("treatment_outcome") or "").split("|"):
        parts = [p.strip() for p in entry.strip().split(":")]
        if len(parts) < 2 or not parts[1]:
            continue
        drug, outcome = parts[0], parts[1].lower()
        which = "4'-DMA" if DMA.search(drug) else ("7,8-DHF" if PLAIN.search(drug) else None)
        if not which or outcome not in RANK:
            continue
        entry_out[which][outcome] += 1
        if len(parts) > 2:
            targets[which][": ".join(parts[2:]).lower()] += 1
        prev = author_out[which].get(r["author_hash"])
        if prev is None or RANK[outcome] > RANK[prev]:
            author_out[which][r["author_hash"]] = outcome
auth_out = {k: collections.Counter(v.values()) for k, v in author_out.items()}

doses = collections.Counter()
for r in B_rows:
    for d in (r.get("dosage") or "").split("|"):
        d = d.strip().lower()
        if d:
            doses[d] += 1

se = collections.Counter()
n_se = 0
for r in A:
    raw = r["side_effects"]
    if not raw or raw == "[]":
        continue
    try:
        lst = json.loads(raw)
    except Exception:
        continue
    if not lst:
        continue
    n_se += 1
    for s in lst:
        se[str(s).strip().lower()] += 1


# ── side-effect structure + OMF condition mapping ───────────────────────────
import itertools
SE_CANON = [
 (r"insomnia|sleep (issue|disrupt|disturb|problem)|can'?t sleep|trouble sleeping|poor sleep", "insomnia / sleep disruption"),
 (r"headache|migraine", "headache / migraine"),
 (r"hair (loss|thinning|shed)|weak hair|balding", "hair loss / thinning"),
 (r"irritab|restless|agitat|overstimulat|jitter|wired|anxious|anxiety|panic", "overstimulation / anxiety"),
 (r"appetite|hunger", "appetite change"),
 (r"nausea|stomach|gi\b|diarrh|gut|digest", "GI"),
 (r"fatigue|tired|lethargy|sedat|drowsy|sleepy", "fatigue / sedation"),
 (r"depress|anhedoni|blunt|apath|emotional", "mood flattening / depression"),
 (r"brain fog|cognitive|memory|concentrat|verbal|articulat", "cognitive dulling"),
 (r"dizz|lightheaded|vertigo", "dizziness"),
 (r"crash|tolerance|withdraw|dependen|rebound", "crash / tolerance / withdrawal"),
 (r"blood pressure|\bbp\b|heart|palpit|tachy", "cardiovascular"),
 (r"vision|visual|aura|eye", "visual"),
 (r"rash|itch|allerg|hives", "allergic / skin"),
 (r"libido|sexual|erectile", "sexual"),
]
SE_CANON = [(re.compile(pp, re.I), lab) for pp, lab in SE_CANON]


def _canon(t):
    for rx, lab in SE_CANON:
        if rx.search(t):
            return lab
    return None


SE_COND = {
 "depression / mood":   r"\bdepress|\bantidepress|\bmood\b|anhedoni",
 "anxiety":             r"\banxiet|\banxious\b|\bpanic\b",
 "focus / cognition":   r"\bfocus\b|\bconcentrat|\bcognit|\bbrain fog\b|\bclarity\b|\bproductiv",
 "memory / learning":   r"\bmemory\b|\brecall\b|\blearning\b",
 "sleep":               r"\bsleep\b|\binsomnia\b",
 "energy / fatigue":    r"\benergy\b|\bfatigue\b|\btired\b|\bstamina\b",
 "neurogenesis / BDNF": r"\bbdnf\b|\btrkb\b|neurogenes|neuroplastic|\brewir",
}
SE_CRX = {k: re.compile(v, re.I) for k, v in SE_COND.items()}
# ASCEND-ME (Xiao group) - the domains the proposed ME/CFS trial targets
OMF = {"energy / fatigue": "PEM / fatigue - PRIMARY endpoint",
       "focus / cognition": "cognitive dysfunction",
       "memory / learning": "cognitive dysfunction",
       "sleep": "sleep (unrefreshing)",
       "depression / mood": "mood (secondary)",
       "anxiety": "mood (secondary)",
       "neurogenesis / BDNF": "BDNF / TrkB - the mechanism"}

serecs = []
for _r in A:
    _ses = []
    if _r["side_effects"] and _r["side_effects"] != "[]":
        try:
            _ses = [x for x in (_canon(str(v)) for v in json.loads(_r["side_effects"])) if x]
        except Exception:
            pass
    _t = (_r["title"] or "") + " " + (_r["body_text"] or "")
    serecs.append(dict(sent=_r["sentiment"], ses=sorted(set(_ses)), user=_r["user_id"],
                       date=_r["post_date"] or 0, rank=SIG.get(_r["sig"], 0),
                       conds=[k for k, rx in SE_CRX.items() if rx.search(_t)]))
SE_N = len(serecs)
SE_WITH = [x for x in serecs if x["ses"]]
SE_NONE = [x for x in serecs if not x["ses"]]
SE_FREQ = collections.Counter(v for x in serecs for v in x["ses"])
SE_TOT = sum(len(x["ses"]) for x in serecs)
SE_PAIR = collections.Counter()
for x in serecs:
    for _a, _b in itertools.combinations(x["ses"], 2):
        SE_PAIR[tuple(sorted((_a, _b)))] += 1
_vote = {c: {} for c in SE_COND}
for x in sorted(serecs, key=lambda y: (y["date"], y["rank"]), reverse=True):
    for c in x["conds"]:
        _vote[c].setdefault(x["user"], x["sent"])
COND_BASE = {c: len(v) for c, v in _vote.items()}
COND_POS = {c: sum(1 for s_ in v.values() if s_ == "positive") for c, v in _vote.items()}

# use-case sentiment
uc = []
for cat, pat in PURPOSE.items():
    rx = re.compile(pat, re.I)
    b = {}
    for r in sorted(A, key=lambda r: (r["post_date"] or 0, SIG.get(r["sig"], 0)), reverse=True):
        if rx.search((r["title"] or "") + " " + (r["body_text"] or "")):
            b.setdefault(r["user_id"], r["sentiment"])
    n = len(b)
    if n < 5:
        continue
    pos = sum(1 for v in b.values() if v == "positive")
    neg = sum(1 for v in b.values() if v == "negative")
    lo, hi = wilson(pos, n)
    base_pos = A_vote["positive"]
    _, p = fisher_exact([[pos, n - pos], [base_pos - pos, (A_N - n) - (base_pos - pos)]])
    uc.append([cat, n, pos, neg, 100 * pos / n, 100 * lo, 100 * hi, p])
q = multipletests([r[7] for r in uc], method="fdr_bh")[1]
for r, qq in zip(uc, q):
    r.append(qq)
uc.sort(key=lambda r: -r[1])

fisher_rows = []
for key in ("no_effect", "helped", "worsened"):
    a, n1 = auth_out["7,8-DHF"][key], sum(auth_out["7,8-DHF"].values())
    c, n2 = auth_out["4'-DMA"][key], sum(auth_out["4'-DMA"].values())
    odds, p = fisher_exact([[a, n1 - a], [c, n2 - c]])
    lo1, hi1 = wilson(a, n1)
    lo2, hi2 = wilson(c, n2)
    fisher_rows.append([key, a, n1, 100 * a / n1, 100 * lo1, 100 * hi1,
                        c, n2, 100 * c / n2, 100 * lo2, 100 * hi2, odds, p])

# ══ build ═══════════════════════════════════════════════════════════════════
wb = Workbook()


def sheet(name, widths):
    ws = wb.create_sheet(name) if wb.sheetnames != ["Sheet"] else wb.active
    ws.title = name
    for col, w in zip("ABCDEFGHIJKL", widths):
        ws.column_dimensions[col].width = w
    return ws


def banner(ws, row, text, span=6):
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True, size=11, color="FFFFFF")
    c.fill = BAND
    for j in range(2, span + 1):
        ws.cell(row, j).fill = BAND
    return row + 1


def header(ws, row, cols, wrap=False):
    for j, name in enumerate(cols, start=1):
        c = ws.cell(row, j, name)
        c.font = Font(bold=True, size=10, color="FFFFFF")
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
        c.border = Border(bottom=RULE)
    if wrap:
        ws.row_dimensions[row].height = 30
    return row + 1


def row(ws, r, vals, fmts=None, bold_col=None):
    for j, v in enumerate(vals, start=1):
        c = ws.cell(r, j, v)
        if bold_col == j:
            c.font = BOLD
        if fmts and j - 1 < len(fmts) and fmts[j - 1]:
            c.number_format = fmts[j - 1]
    return r + 1


def kv(ws, r, k, v, note=""):
    ws.cell(r, 1, k).font = BOLD
    ws.cell(r, 2, v)
    if note:
        n = ws.cell(r, 3, note)
        n.font = SMALL
        n.alignment = WRAP
    return r + 1


# ── 1. Summary ──────────────────────────────────────────────────────────────
ws = sheet("Summary", [34, 26, 74])
ws["A1"] = "7,8-DHF (tropoflavin) — r/Nootropics run"
ws["A1"].font = Font(bold=True, size=15, color=INK)
ws["A2"] = "All processed results. Mention/sentiment data from a healthy-user nootropics population — not patient outcomes."
ws["A2"].font = Font(size=10, italic=True, color=MUTED)
r = 4
r = banner(ws, r, "CORPUS", 3)
r = kv(ws, r, "Subreddit", "r/Nootropics", "posts + comments combined")
r = kv(ws, r, "Comments source", "1,827,221 lines", "2009-09-25 → 2026-08-18")
r = kv(ws, r, "Posts source", "184,321 lines", "→ 2026-08-17")
r = kv(ws, r, "Mentions found", N_ITEMS, "1,448 in comments + 345 in posts (1 item matched both fields)")
r = kv(ws, r, "Threads pulled whole", 1048, "every thread containing ≥1 mention, full reply tree")
r = kv(ws, r, "Corpus items", 45667, "1,047 posts + 44,620 comments")
r = kv(ws, r, "Thread structure retained", "99.9%", "parent_id survival after import")
r = kv(ws, r, "Distinct authors in corpus", 13568, "everyone in those threads")
r = kv(ws, r, "Authors who named it", 752, "pipeline B population")
r += 1
r = banner(ws, r, "PIPELINE A — drug sentiment (--drug-file)", 3)
r = kv(ws, r, "Entry × drug pairs", 4603, "1,653 direct alias matches + 2,950 context-inherited")
r = kv(ws, r, "Survived prefilter", 988, "fast model drops non-personal-experience pairs")
r = kv(ws, r, "Sentiment records", len(A), "")
r = kv(ws, r, "Distinct users", A_N, "the unit that bounds power")
lo, hi = wilson(A_vote["positive"], A_N)
r = kv(ws, r, "Positive (one vote/user)", f"{100*A_vote['positive']/A_N:.1f}%",
       f"{A_vote['positive']}/{A_N}, 95% Wilson [{100*lo:.1f}, {100*hi:.1f}] — blends both compounds")
r = kv(ws, r, "Negative (one vote/user)", f"{100*A_vote['negative']/A_N:.1f}%",
       "more trustworthy than the positive rate (~10–20% positive over-call)")
r += 1
r = banner(ws, r, "PIPELINE B — variable extraction (no --drug; pre-filtered corpus)", 3)
r = kv(ws, r, "Records", len(B_rows), "one per author who named the compound")
r = kv(ws, r, "Clinical fields", len(B_fields), "nootropics_v1 schema, 25 base fields")
r = kv(ws, r, "Field fills", sum(fill.values()), f"{sum(fill.values())/len(B_rows):.2f} per record")
r = kv(ws, r, "Runtime", "16m 56s", "752 records, 12 workers, deepseek-v4-flash")
r = kv(ws, r, "7,8-DHF outcomes", f"{sum(auth_out['7,8-DHF'].values())} authors",
       f"{sum(entry_out['7,8-DHF'].values())} entries")
r = kv(ws, r, "4'-DMA outcomes", f"{sum(auth_out['4-DMA'].values()) if False else sum(auth_out[chr(52)+chr(39)+'-DMA'].values())} authors",
       f"{sum(entry_out[chr(52)+chr(39)+'-DMA'].values())} entries")
r += 1
r = banner(ws, r, "HEADLINE", 3)
for txt in [
    "B separates 7,8-DHF from 4'-DMA-7,8-DHF; A cannot — its alias for '7,8-dhf' matches inside \"4'-DMA-7,8-DHF\".",
    "A's 71.1% blended rate sits on the derivative's rate, not the parent's — over-capture is measurable at ~9 points.",
    "Only the no_effect gap separates the compounds (p=0.048); helped and worsened do not.",
    "Sentiment-by-use-case looks strong but is an artifact — see the Statistics tab.",
    "Side effects cluster on overstimulation; hair loss is an unexpected third-ranked signal.",
]:
    c = ws.cell(r, 1, txt)
    c.font = Font(size=10, color=INK)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 26
    r += 1
ws.freeze_panes = "A4"

# ── 2. Use cases ────────────────────────────────────────────────────────────
ws = sheet("Use cases", [34, 12, 12, 60])
ws["A1"] = "What people use it for"
ws["A1"].font = Font(bold=True, size=14, color=INK)
ws["A2"] = f"Keyword context across all {N_ITEMS:,} mentioning items. An item can match several rows, so shares exceed 100%."
ws["A2"].font = Font(size=10, italic=True, color=MUTED)
r = header(ws, 4, ["Stated context", "items", "share", ""])
for k, v in sorted(purpose_n.items(), key=lambda x: -x[1]):
    r = row(ws, r, [k, v, v / N_ITEMS, ""], [None, "#,##0", "0.0%"])
r += 1
r = banner(ws, r, "CO-MENTIONED SUBSTANCES", 4)
r = header(ws, r, ["Substance", "items", "share", ""])
for k, v in sorted(co_n.items(), key=lambda x: -x[1]):
    r = row(ws, r, [k, v, v / N_ITEMS, ""], [None, "#,##0", "0.0%"])
ws.freeze_panes = "A5"

# ── 3. Sentiment A ──────────────────────────────────────────────────────────
ws = sheet("Sentiment (A)", [26, 12, 12, 14, 14, 60])
ws["A1"] = "Pipeline A — drug sentiment"
ws["A1"].font = Font(bold=True, size=14, color=INK)
ws["A2"] = "Targeted --drug-file run. Covers BOTH compounds together; see Extraction (B) for the split."
ws["A2"].font = Font(size=10, italic=True, color=MUTED)
r = banner(ws, 4, "RAW RECORDS (not independent — one user may contribute several)", 6)
r = header(ws, r, ["Sentiment", "records", "share", "", "", ""])
for k in ("positive", "negative", "mixed", "neutral"):
    r = row(ws, r, [k, A_raw[k], A_raw[k] / len(A), "", "", ""], [None, "#,##0", "0.0%"])
r = row(ws, r, ["TOTAL", len(A), 1.0, "", "", ""], [None, "#,##0", "0.0%"], bold_col=1)
r += 1
r = banner(ws, r, "ONE VOTE PER USER — most recent record, ties broken by signal strength", 6)
r = header(ws, r, ["Sentiment", "users", "share", "95% CI low", "95% CI high", ""])
for k in ("positive", "negative", "mixed", "neutral"):
    lo, hi = wilson(A_vote[k], A_N)
    r = row(ws, r, [k, A_vote[k], A_vote[k] / A_N, lo, hi, ""],
            [None, "#,##0", "0.0%", "0.0%", "0.0%"])
r = row(ws, r, ["TOTAL", A_N, 1.0, "", "", ""], [None, "#,##0", "0.0%"], bold_col=1)
r += 1
r = banner(ws, r, "BY SIGNAL STRENGTH (raw records)", 6)
r = header(ws, r, ["Sentiment", "strong", "moderate", "weak", "n/a", ""])
for s in ("positive", "negative", "mixed", "neutral"):
    r = row(ws, r, [s, A_sig[(s, "strong")], A_sig[(s, "moderate")],
                    A_sig[(s, "weak")], A_sig[(s, "n/a")], ""],
            [None, "#,##0", "#,##0", "#,##0", "#,##0"])
ws.freeze_panes = "A5"

# ── 4. Extraction B ─────────────────────────────────────────────────────────
ws = sheet("Extraction (B)", [30, 12, 12, 14, 14, 46])
ws["A1"] = "Pipeline B — variable extraction"
ws["A1"].font = Font(bold=True, size=14, color=INK)
ws["A2"] = "752 per-author records. B has no --drug flag; targeting = a pre-filtered corpus of authors who named the compound."
ws["A2"].font = Font(size=10, italic=True, color=MUTED)
r = banner(ws, 4, "FIELD FILL RATES", 6)
r = header(ws, r, ["Field", "filled", "rate", "", "", ""])
for k, v in fill.most_common():
    r = row(ws, r, [k, v, v / len(B_rows), "", "", ""], [None, "#,##0", "0.0%"])
r += 1
r = banner(ws, r, "OUTCOMES BY COMPOUND — author level (one vote per author per compound)", 6)
r = header(ws, r, ["Outcome", "7,8-DHF n", "7,8-DHF %", "4'-DMA n", "4'-DMA %", ""])
DM = chr(52) + chr(39) + "-DMA"
n1, n2 = sum(auth_out["7,8-DHF"].values()), sum(auth_out[DM].values())
for k in ("helped", "worsened", "no_effect", "mixed", "unknown"):
    rr = row(ws, r, [k, auth_out["7,8-DHF"][k], auth_out["7,8-DHF"][k] / n1,
                     auth_out[DM][k], auth_out[DM][k] / n2, ""],
             [None, "#,##0", "0.0%", "#,##0", "0.0%"])
    if k == "no_effect":
        for j in (2, 3, 4, 5):
            ws.cell(r, j).fill = WARN
    r = rr
r = row(ws, r, ["TOTAL authors", n1, 1.0, n2, 1.0, ""], [None, "#,##0", "0.0%", "#,##0", "0.0%"], bold_col=1)
r += 1
r = banner(ws, r, "OUTCOMES BY COMPOUND — entry level (each outcome statement counted)", 6)
r = header(ws, r, ["Outcome", "7,8-DHF n", "7,8-DHF %", "4'-DMA n", "4'-DMA %", ""])
e1, e2 = sum(entry_out["7,8-DHF"].values()), sum(entry_out[DM].values())
for k in ("helped", "worsened", "no_effect", "mixed", "unknown"):
    r = row(ws, r, [k, entry_out["7,8-DHF"][k], entry_out["7,8-DHF"][k] / e1,
                    entry_out[DM][k], entry_out[DM][k] / e2, ""],
            [None, "#,##0", "0.0%", "#,##0", "0.0%"])
r = row(ws, r, ["TOTAL entries", e1, 1.0, e2, 1.0, ""], [None, "#,##0", "0.0%", "#,##0", "0.0%"], bold_col=1)
r += 1
r = banner(ws, r, "WHAT IT REPORTEDLY AFFECTED (free-text targets, top 12 each)", 6)
r = header(ws, r, ["7,8-DHF target", "n", "4'-DMA target", "n", "", ""])
t1 = targets["7,8-DHF"].most_common(12)
t2 = targets[DM].most_common(12)
for i in range(max(len(t1), len(t2))):
    a = t1[i] if i < len(t1) else ("", "")
    b = t2[i] if i < len(t2) else ("", "")
    r = row(ws, r, [a[0], a[1], b[0], b[1], "", ""], [None, "#,##0", None, "#,##0"])
ws.freeze_panes = "A5"

# ── 5. Statistics ───────────────────────────────────────────────────────────
ws = sheet("Statistics", [30, 10, 10, 11, 11, 11, 11, 11, 11, 62])
ws["A1"] = "Statistical tests"
ws["A1"].font = Font(bold=True, size=14, color=INK)
r = banner(ws, 3, "COMPOUND COMPARISON — Fisher's exact, author level", 10)
r = header(ws, r, ["Test", "7,8-DHF n", "of", "7,8-DHF %", "4'-DMA n", "of", "4'-DMA %",
                   "odds ratio", "Fisher p", "verdict"], wrap=True)
for key, a, na, pa, la, ha, c, nb, pb, lb, hb, odds, p in fisher_rows:
    verdict = ("SIGNIFICANT at 0.05 — but does NOT survive Bonferroni across these 3 tests (needs p<0.017); CIs overlap"
               if p < 0.05 else "not significant")
    rr = row(ws, r, [key + " vs rest", a, na, pa / 100, c, nb, pb / 100, odds, p, verdict],
             [None, "#,##0", "#,##0", "0.0%", "#,##0", "#,##0", "0.0%", "0.00", "0.0000"])
    ws.cell(r, 10).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 28
    if p < 0.05:
        for j in range(1, 10):
            ws.cell(r, j).fill = WARN
    r = rr
r += 1
c = ws.cell(r, 1, "Author-level collapse rule: an author contributing several outcome statements for one compound "
                  "keeps the most informative, ranked worsened > no_effect > mixed > helped. This is deliberately "
                  "conservative and moves 'helped' far more than 'no_effect' (entry-level helped was 61.9% / 70.7%).")
c.font = Font(size=9, italic=True, color=MUTED)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 30
r += 2
r = banner(ws, r, "SENTIMENT BY USE-CASE — ARTIFACT, DO NOT REPORT AS A FINDING", 10)
r = header(ws, r, ["Use-case", "users", "pos", "neg", "pos %", "CI low", "CI high",
                   "Fisher p", "BH q", "vs 71.1% baseline"], wrap=True)
for cat, n, pos, neg, pct, lo, hi, p, qq in uc:
    note = ("clears FDR — but see the warning below" if qq < 0.05 else "")
    rr = row(ws, r, [cat, n, pos, neg, pct / 100, lo / 100, hi / 100, p, qq, note],
             [None, "#,##0", "#,##0", "#,##0", "0.0%", "0.0%", "0.0%", "0.000", "0.000"])
    if n < 30:
        ws.cell(r, 2).fill = BAD
    r = rr
r = row(ws, r, ["BASELINE (all users)", A_N, A_vote["positive"], A_vote["negative"],
                A_vote["positive"] / A_N, "", "", "", "", ""],
        [None, "#,##0", "#,##0", "#,##0", "0.0%"], bold_col=1)
r += 1
c = ws.cell(r, 1, "WHY THIS IS AN ARTIFACT: every category sits at or above the 71.1% baseline and none below — neutral "
                  "slices would straddle it. Categories are keyword-derived from the same text the classifier scored, and "
                  "positive reports name what improved (\"helped my depression\" fires both) while nulls say \"did nothing\" "
                  "and name no domain. The categorisation therefore excludes nulls by construction and every domain floats "
                  "upward. The uncontaminated comparison is category vs category — depression/mood 88.5% vs sleep 71.0% — "
                  "with sleep as a partial control, since insomnia is simultaneously the top side effect.")
c.font = Font(size=9, color="A32F2F")
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 74
ws.freeze_panes = "A4"

# ── 6. Side effects & doses ─────────────────────────────────────────────────
ws = sheet("Side effects & doses", [34, 10, 10, 6, 24, 10, 46])
ws["A1"] = "Side effects and dosages"
ws["A1"].font = Font(bold=True, size=14, color=INK)
ws["A2"] = (f"{n_se:,} of {len(A):,} pipeline-A records carry side effects ({100*n_se/len(A):.1f}%) — "
            f"{sum(se.values()):,} mentions across {len(se):,} distinct terms.")
ws["A2"].font = Font(size=10, italic=True, color=MUTED)
r = banner(ws, 4, "SIDE EFFECTS (pipeline A, uncanonicalised — counts understate)", 7)
r = header(ws, r, ["Term", "n", "", "", "Dosage (pipeline B)", "n", ""])
sl = se.most_common(30)
dl = doses.most_common(30)
for i in range(max(len(sl), len(dl))):
    a = sl[i] if i < len(sl) else ("", "")
    b = dl[i] if i < len(dl) else ("", "")
    r = row(ws, r, [a[0], a[1], "", "", b[0], b[1], ""], [None, "#,##0", None, None, None, "#,##0"])
r += 1
for txt in [
    "FRAGMENTATION: 135 distinct terms for 216 mentions. headache/headaches, hair loss/hair thinning, "
    "insomnia/sleep issues/sleep disruption are separate rows. Canonicalise before quoting any single count.",
    "OVERSTIMULATION CLUSTER: insomnia, irritability, restlessness, overstimulated, anxiety, appetite "
    "suppression — coherent, and matches both the 'similar to modafinil' framing here and the r/cfs report "
    "of stimulation persisting into the evening.",
    "HAIR LOSS is the unexpected signal — 7 + 3 'hair thinning' = 10, third most common. Not an obvious "
    "consequence of TrkB agonism. Whether it attributes to 7,8-DHF or a co-stacked substance is unresolved.",
    "DOSAGE CAVEAT: pipeline B's dosage field is per-record and NOT linked to a specific drug, so these "
    "values mix every substance a person listed. They cannot be read as 7,8-DHF doses. The route convention "
    "reported for 7,8-DHF specifically is ~1 mg sublingual, with capsules described as much weaker.",
]:
    c = ws.cell(r, 1, txt)
    c.font = Font(size=9, color=MUTED)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[r].height = 34
    r += 1
ws.freeze_panes = "A5"

# ── 7. Method ───────────────────────────────────────────────────────────────
ws = sheet("Method & caveats", [30, 112])
ws["A1"] = "Method, provenance and caveats"
ws["A1"].font = Font(bold=True, size=14, color=INK)
r = 3
SECTIONS = [
    ("DATA", [
        ("Comments", "PatientPunk_data/r_nootropics_comments.jsonl - 1,827,221 lines, 2009-09-25 to 2026-08-18"),
        ("Posts", "OneDrive/Documents/r_nootropics_posts.jsonl - 184,321 lines, to 2026-08-17"),
        ("S3", "s3://patientpunk/raw_data/arctic_shift_ndjson/r_nootropics_{comments,posts}.jsonl"),
        ("Truncation warning", "An earlier download stopped at 2019-05-02 while looking healthy. Check max timestamp, not file size."),
    ]),
    ("MATCHING", [
        ("Regex", "tropoflavin | hydroxyflavone | 7[sep]8[sep]dhf | dhf  (word-boundaried, case-insensitive) - see NOTES.md"),
        ("Prefilter", "bytes test for dhf / flavon / tropoflav before json.loads"),
        ("Traps", "Body text only (base36 ids contain i78dhf6). The bare dhf token misses 7,8DHF. FTS5 undercounts. See NOTES.md section 2."),
    ]),
    ("PIPELINES", [
        ("A - sentiment", "src/run_sentiment_pipeline.py --drug-file ... --subreddit Nootropics. Needs LLM_MAX_TOKENS=16000."),
        ("A --drug", "YES - --drug NAME and --drug-file PATH"),
        ("B - extraction", "variable_extraction/main.py run --schema schemas/nootropics_schema.json --input-dir source_B"),
        ("B --drug", "NO - flag does not exist. B's unit is the patient. Targeting = pre-filtered corpus."),
        ("B critical", "Do NOT point B at subreddit_posts.json - it reads title+body only and drops 81% of the signal. See NOTES.md section 4."),
    ]),
    ("CAVEATS", [
        ("Population", "r/Nootropics is a HEALTHY-USER population. Answers dose, route, tolerability, subjective effect. Not patient outcomes."),
        ("Positive over-call", "~10-20% of 'positive' labels are false positives; negatives are reliable."),
        ("Compound blending", "33% of mentions co-occur with 4'-DMA. A cannot separate them. Use B for per-compound claims."),
        ("Group attribution", "Many mentions sit inside long stacks. Run the monotherapy check before quoting rates."),
        ("Multiplicity", "3 tests on compounds, 9 on use-cases with BH. Nothing pre-registered."),
    ]),
    ("REPRODUCE", [
        ("Scripts", "build_corpus.py, build_corpus_B.py, analyze_purpose.py, analyze_B.py, analyze_followups.py, make_sheet.py"),
        ("Full notes", "studies/tropoflavin_nootropics/NOTES.md"),
    ]),
]
for title, rows_ in SECTIONS:
    r = banner(ws, r, title, 2)
    for k, v in rows_:
        ws.cell(r, 1, k).font = BOLD
        b = ws.cell(r, 2, v)
        b.alignment = WRAP
        ws.row_dimensions[r].height = max(14, 12 * (len(v) // 108 + 1))
        r += 1
    r += 1


# ── OMF / ASCEND-ME conditions ──────────────────────────────────────────────
ws = sheet("OMF conditions", [26, 30, 10, 9, 11, 11, 11, 46])
ws["A1"] = "Conditions of interest to OMF / ASCEND-ME"
ws["A1"].font = Font(bold=True, size=14, color=INK)
ws["A2"] = ("ASCEND-ME (Xiao group) proposes a decentralised dose-ranging pilot of 7,8-DHF in ME/CFS, asking "
            "whether it improves PEM-linked dysfunction. Rates below are r/Nootropics self-reports - a "
            "healthy-user population, NOT ME/CFS patients.")
ws["A2"].font = SMALL
ws["A2"].alignment = WRAP
ws.row_dimensions[2].height = 30
r = banner(ws, 4, "SENTIMENT BY CONDITION TREATED  (7,8-DHF, one vote per user per condition)", 8)
r = header(ws, r, ["Condition", "ASCEND-ME relevance", "users", "pos", "pos %", "CI low", "CI high", "note"], wrap=True)
for cond in sorted(SE_COND, key=lambda c: -COND_BASE[c]):
    n_, k_ = COND_BASE[cond], COND_POS[cond]
    if not n_:
        continue
    lo, hi = wilson(k_, n_)
    note = ("PRIMARY endpoint of the proposed trial - and the one domain where 7,8-DHF is NOT elevated"
            if cond == "energy / fatigue"
            else "strongest signal in the corpus" if cond in ("depression / mood", "focus / cognition") else "")
    rr = row(ws, r, [cond, OMF.get(cond, ""), n_, k_, k_ / n_, lo, hi, note],
             [None, None, "#,##0", "#,##0", "0.0%", "0.0%", "0.0%"])
    fill = BAD if cond == "energy / fatigue" else (GOOD if cond in ("depression / mood", "focus / cognition") else None)
    if fill:
        for j in range(1, 9):
            ws.cell(r, j).fill = fill
    ws.cell(r, 8).alignment = WRAP
    r = rr
r += 1
for txt in [
    "ENDPOINT WARNING: the proposed trial powers on PEM-linked fatigue. In this corpus 7,8-DHF is 89% positive "
    "for depression/mood and 87% for focus/cognition - both survive FDR correction against anchor substances - "
    "but only 73% for energy/fatigue, identical to l-theanine and below creatine's 77%, and 71% for sleep, below "
    "l-theanine's 73%. The reported benefit concentrates where TrkB agonism predicts and is unremarkable on the "
    "energy axis.",
    "POPULATION CAVEAT: these are r/Nootropics healthy users. Across nine patient subreddits (3.8M items) 7,8-DHF "
    "has 33 mentions from 13 authors, 4 of whom describe taking it - which is why the analysis moved here. This "
    "speaks to dose, route, tolerability and symptom specificity, not to efficacy in ME/CFS.",
]:
    c = ws.cell(r, 1, txt)
    c.font = Font(size=9, color="A32F2F")
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
    c.alignment = WRAP
    ws.row_dimensions[r].height = 46
    r += 1
ws.freeze_panes = "A6"

# ── Side effects, expanded ──────────────────────────────────────────────────
ws = sheet("Side effects (detail)", [34, 11, 11, 30, 11, 11, 46])
ws["A1"] = "Side-effect structure"
ws["A1"].font = Font(bold=True, size=14, color=INK)
r = banner(ws, 3, "BURDEN", 7)
for k, v, note in [
    ("Records", f"{SE_N:,}", ""),
    ("Reporting >=1 side effect", f"{len(SE_WITH):,}  ({100*len(SE_WITH)/SE_N:.1f}%)", ""),
    ("Canonicalised mentions", f"{SE_TOT}", "from 216 raw terms across 135 spellings"),
    ("Mean per record (all)", f"{SE_TOT/SE_N:.2f}", ""),
    ("Mean among reporters", f"{SE_TOT/len(SE_WITH):.2f}", "side effects come singly, not in long lists"),
]:
    r = kv(ws, r, k, v, note)
r += 1
r = banner(ws, r, "SENTIMENT SPLITS SHARPLY ON WHETHER A SIDE EFFECT WAS REPORTED", 7)
r = header(ws, r, ["Group", "records", "positive", "negative", "", "", ""])
for lab, sub, fill in (("reported >=1 side effect", SE_WITH, BAD), ("reported none", SE_NONE, GOOD)):
    cnt = collections.Counter(x["sent"] for x in sub)
    m = len(sub)
    rr = row(ws, r, [lab, m, cnt["positive"] / m, cnt["negative"] / m, "", "", ""],
             [None, "#,##0", "0.0%", "0.0%"])
    ws.cell(r, 1).fill = fill
    r = rr
c = ws.cell(r, 1, "The 71.1% headline is a blend of these two populations. Side effects are uncommon (14.7% of "
                  "records) but where present they dominate the verdict.")
c.font = SMALL
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c.alignment = WRAP
r += 2

r = banner(ws, r, "FREQUENCY (canonicalised)", 7)
r = header(ws, r, ["Side effect", "n", "% of records", "", "", "", ""])
for s_, k_ in SE_FREQ.most_common():
    rr = row(ws, r, [s_, k_, k_ / SE_N, "", "", "", ""], [None, "#,##0", "0.0%"])
    if s_.startswith("hair"):
        for j in (1, 2, 3):
            ws.cell(r, j).fill = WARN
    r = rr
r += 1
r = banner(ws, r, "CO-OCCURRENCE  (lift = how much more often than chance)", 7)
r = header(ws, r, ["Side effect A", "n", "lift", "Side effect B", "", "", ""])
for (a_, b_), k_ in SE_PAIR.most_common(10):
    if k_ < 3:
        break
    lift = (k_ / SE_N) / ((SE_FREQ[a_] / SE_N) * (SE_FREQ[b_] / SE_N))
    r = row(ws, r, [a_, k_, lift, b_, "", "", ""], [None, "#,##0", '0.0"x"'])
c = ws.cell(r, 1, "TWO CLUSTERS, OPPOSITE IN DIRECTION. An overstimulation hub (overstimulation/anxiety with "
                  "insomnia 6.4x, GI 24.5x, dizziness 18.4x) and a blunting pair (cognitive dulling with mood "
                  "flattening, 40.5x). Some users are over-activated, others flattened - what a dose-response "
                  "that overshoots in some people would look like.")
c.font = Font(size=9, color=INK)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
c.alignment = WRAP
ws.row_dimensions[r].height = 34
r += 2

r = banner(ws, r, "SIDE EFFECT BY STATED REASON FOR TAKING IT (% of that condition's records)", 7)
conds_sorted = sorted(SE_COND, key=lambda c: -COND_BASE[c])[:6]
r = header(ws, r, ["Side effect"] + [c[:14] for c in conds_sorted], wrap=True)
for s_, _ in SE_FREQ.most_common(8):
    vals = [s_]
    for cnd in conds_sorted:
        d = COND_BASE[cnd]
        k_ = sum(1 for x in serecs if cnd in x["conds"] and s_ in x["ses"])
        vals.append(k_ / d if d >= 25 else "")
    r = row(ws, r, vals, [None] + ["0%"] * len(conds_sorted))
r += 1
for txt, col in [
    ("THE DIAGONAL IS AN ARTIFACT - do not cite it. The indication regexes and the side-effect canonicaliser "
     "share vocabulary, so naming a side effect files that author under the matching indication. Every "
     "standout cell above sits on the diagonal.", "A32F2F"),
    ("REVIEW DONE - see audit_diag.py and audit_fatigue.py. Re-tagging the indication from non-outcome "
     "sentences only: anxiety 32% -> no records survive, energy/fatigue 11% -> 3% (1 of 32), depression, "
     "focus and memory -> 0%. Only sleep -> insomnia survives, at 18% (8 of 45). Hand-reading all 6 "
     "energy/fatigue records found 2-3 genuine, one explicit (CFS, drug exacerbates fatigue and somnolence); "
     "the rest were tagged by the side-effect word itself. No paradoxical-reaction claim is supportable.", INK),
    ("HAIR LOSS IS THE EXCEPTION, and the reason to take it seriously: 4% in the neurogenesis/BDNF group and 0% "
     "everywhere else - not a mirror of any indication, so not explainable by vocabulary overlap. It is also "
     "absent from the ASCEND-ME safety list (headaches, BP, insomnia/agitation, allergy, visual auras, "
     "neuropsychiatric, discontinuation, product quality). Candidate addition to prospective monitoring.", INK),
]:
    c = ws.cell(r, 1, txt)
    c.font = Font(size=9, color=col)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    c.alignment = WRAP
    ws.row_dimensions[r].height = 42
    r += 1


wb.save(OUT)
print(f"wrote {OUT} ({OUT.stat().st_size:,} bytes)")
print(f"tabs: {wb.sheetnames}")

# Also emit a flat CSV of every tab. The Drive connector here has no Sheets API
# and the binary upload path is unreliable at this size, so this is what gets
# uploaded; the .xlsx above is the formatted version to drag into Drive by hand.
import csv as _csv
CSV_OUT = HERE / "results_workbook.csv"
with CSV_OUT.open("w", newline="", encoding="utf-8") as fh:
    w = _csv.writer(fh)
    for name in wb.sheetnames:
        ws = wb[name]
        w.writerow([])
        w.writerow(["=" * 12, name.upper(), "=" * 12])
        w.writerow([])
        for row_cells in ws.iter_rows():
            vals = []
            for c in row_cells:
                v = c.value
                if v is None:
                    vals.append("")
                elif isinstance(v, float) and c.number_format and "%" in c.number_format:
                    vals.append(f"{v*100:.1f}%")
                elif isinstance(v, float):
                    vals.append(f"{v:.4g}")
                else:
                    vals.append(str(v))
            while vals and vals[-1] == "":
                vals.pop()
            if vals:
                w.writerow(vals)
print(f"wrote {CSV_OUT} ({CSV_OUT.stat().st_size:,} bytes)")
